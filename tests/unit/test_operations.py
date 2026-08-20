from __future__ import annotations

import logging
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook
from typer.testing import CliRunner

from research_gateway.cli import app
from research_gateway.config import Settings
from research_gateway.db.database import EvidenceDatabase
from research_gateway.operations.backups import ExcelBackupService
from research_gateway.operations.logging import configure_logging
from research_gateway.operations.service import ServiceManager, ServiceStartError
from research_gateway.operations.supervisor import SystemdUserSupervisor


@pytest.mark.asyncio
async def test_excel_backup_creates_timestamped_and_latest_workbooks(tmp_path: Path) -> None:
    database = EvidenceDatabase(tmp_path / "data" / "research_gateway.db")
    await database.migrate()
    service = ExcelBackupService(database, tmp_path / "backups", retention_count=2)

    result = await service.create()

    assert result.path.is_file()
    assert result.latest_path.is_file()
    assert result.path != result.latest_path
    assert "Evidence" in load_workbook(result.latest_path, read_only=True).sheetnames


def test_file_logging_redacts_all_configured_secrets(tmp_path: Path) -> None:
    settings = Settings.model_validate(
        {
            "database": {"path": tmp_path / "data" / "research_gateway.db"},
            "logging": {"path": tmp_path / "logs" / "research-gateway.log"},
            "scopus": {"api_key": "fixture-secret-value"},
        }
    )
    log_path = configure_logging(settings)
    logging.getLogger("research_gateway.test").warning(
        "credential=%s", settings.scopus.api_key.get_secret_value()
    )
    for handler in logging.getLogger().handlers:
        handler.flush()
    assert log_path.is_file()
    assert "fixture-secret-value" not in log_path.read_text(encoding="utf-8")


def test_access_logging_keeps_path_and_status_but_removes_oauth_query(tmp_path: Path) -> None:
    settings = Settings.model_validate(
        {
            "database": {"path": tmp_path / "data" / "research_gateway.db"},
            "logging": {"path": tmp_path / "logs" / "research-gateway.log"},
        }
    )
    log_path = configure_logging(settings)
    access_logger = logging.getLogger("uvicorn.access")
    access_logger.setLevel(logging.INFO)
    access_logger.info(
        '%s - "%s %s HTTP/%s" %d',
        "127.0.0.1:1234",
        "GET",
        "/oauth/authorize?request=raw-secret-request-id&state=raw-state",
        "1.1",
        200,
    )
    for handler in logging.getLogger().handlers:
        handler.flush()

    content = log_path.read_text(encoding="utf-8")
    assert "GET /oauth/authorize HTTP/1.1" in content
    assert "raw-secret-request-id" not in content
    assert "raw-state" not in content


def test_http_client_logging_removes_oauth_query_values(tmp_path: Path) -> None:
    settings = Settings.model_validate(
        {
            "database": {"path": tmp_path / "data" / "research_gateway.db"},
            "logging": {"path": tmp_path / "logs" / "research-gateway.log"},
        }
    )
    log_path = configure_logging(settings)
    client_logger = logging.getLogger("httpx._client")
    client_logger.setLevel(logging.INFO)
    client_logger.info(
        "HTTP Request: %s %s %s",
        "GET",
        "https://example.test/authorize?state=raw-state&request=raw-request-id",
        "HTTP/1.1 302 Found",
    )
    for handler in logging.getLogger().handlers:
        handler.flush()

    content = log_path.read_text(encoding="utf-8")
    assert "GET https://example.test/authorize HTTP/1.1 302 Found" in content
    assert "raw-state" not in content
    assert "raw-request-id" not in content


def test_service_stop_does_not_kill_state_owned_by_another_config(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    manager = _service_manager(tmp_path)
    manager.runtime_directory.mkdir(parents=True)
    manager.state_path.write_text(
        '{"pid": 4321, "config_path": "/another/config.toml"}', encoding="utf-8"
    )
    monkeypatch.setattr(manager, "_process_matches", lambda pid: True)
    monkeypatch.setattr(manager, "_health_ok", lambda: False)
    monkeypatch.setattr(manager, "_port_in_use", lambda: False)
    monkeypatch.setattr(manager, "_discover_process", lambda: None)
    monkeypatch.setattr(
        "research_gateway.operations.service.os.kill",
        lambda pid, sent: pytest.fail("a differently configured process must not be killed"),
    )

    result = manager.stop()

    assert result["stopped"] is False
    assert manager.state_path.exists()


def _service_manager(tmp_path: Path) -> ServiceManager:
    settings = Settings.model_validate(
        {
            "database": {"path": tmp_path / "data" / "research_gateway.db"},
            "logging": {"path": tmp_path / "logs" / "research-gateway.log"},
            "runtime": {"directory": tmp_path / "runtime"},
            "service": {"port": 8877},
        }
    )
    return ServiceManager(settings, tmp_path / "config.toml")


def test_service_start_records_validated_process_state(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    manager = _service_manager(tmp_path)
    statuses = iter([{"running": False}, {"running": True, "pid": 4321}])
    monkeypatch.setattr(manager, "status", lambda: next(statuses))
    waited: list[int] = []
    monkeypatch.setattr(manager, "_wait_for_health", waited.append)
    launched: list[tuple[list[str], dict[str, object]]] = []

    def popen(command: list[str], **kwargs: object) -> SimpleNamespace:
        launched.append((command, kwargs))
        return SimpleNamespace(pid=4321)

    monkeypatch.setattr("research_gateway.operations.service.subprocess.Popen", popen)

    result = manager.start(tunnel=True)

    assert result == {"running": True, "pid": 4321, "started": True}
    assert waited == [4321]
    assert launched[0][0][-1] == "--tunnel"
    assert manager._read_state()["pid"] == 4321


def test_service_start_reuses_a_running_process(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    manager = _service_manager(tmp_path)
    monkeypatch.setattr(
        manager,
        "status",
        lambda: {"running": True, "classification": "managed", "pid": 99},
    )
    assert manager.start(tunnel=False) == {
        "running": True,
        "classification": "managed",
        "pid": 99,
        "started": False,
    }


def test_service_start_does_not_duplicate_unmanaged_healthy_gateway(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    manager = _service_manager(tmp_path)
    launched: list[object] = []
    monkeypatch.setattr(manager, "_health_ok", lambda: True)
    monkeypatch.setattr(manager, "_discover_process", lambda: (7654, manager.config_path))
    monkeypatch.setattr(
        "research_gateway.operations.service.subprocess.Popen",
        lambda *args, **kwargs: launched.append((args, kwargs)),
    )

    result = manager.start(tunnel=True)

    assert result["running"] is True
    assert result["classification"] == "unmanaged"
    assert result["pid"] == 7654
    assert result["started"] is False
    assert launched == []
    assert not manager.state_path.exists()


def test_stale_state_and_healthy_gateway_are_preserved_without_duplicate_start(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    manager = _service_manager(tmp_path)
    manager.runtime_directory.mkdir(parents=True)
    original = b'{"pid": 111, "config_path": "/old/config.toml"}'
    manager.state_path.write_bytes(original)
    monkeypatch.setattr(manager, "_process_matches", lambda pid: False)
    monkeypatch.setattr(manager, "_health_ok", lambda: True)
    monkeypatch.setattr(manager, "_discover_process", lambda: (7654, manager.config_path))
    monkeypatch.setattr(
        "research_gateway.operations.service.subprocess.Popen",
        lambda *args, **kwargs: pytest.fail("a second process must not be launched"),
    )

    result = manager.start(tunnel=True)

    assert result["classification"] == "unmanaged"
    assert result["state_stale"] is True
    assert manager.state_path.read_bytes() == original


def test_port_occupied_by_non_gateway_is_controlled_and_does_not_start(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    manager = _service_manager(tmp_path)
    monkeypatch.setattr(manager, "_health_ok", lambda: False)
    monkeypatch.setattr(manager, "_port_in_use", lambda: True)
    monkeypatch.setattr(
        "research_gateway.operations.service.subprocess.Popen",
        lambda *args, **kwargs: pytest.fail("a process must not be launched on an occupied port"),
    )

    with pytest.raises(ServiceStartError, match="Port 8877 is already in use"):
        manager.start(tunnel=True)


def test_service_start_failure_terminates_owned_process(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    manager = _service_manager(tmp_path)
    monkeypatch.setattr(manager, "status", lambda: {"running": False})
    monkeypatch.setattr(
        "research_gateway.operations.service.subprocess.Popen",
        lambda *args, **kwargs: SimpleNamespace(pid=4321),
    )
    monkeypatch.setattr(
        manager,
        "_wait_for_health",
        lambda pid: (_ for _ in ()).throw(RuntimeError("not healthy")),
    )
    monkeypatch.setattr(manager, "_process_matches", lambda pid: True)
    signals: list[tuple[int, int]] = []
    monkeypatch.setattr(
        "research_gateway.operations.service.os.kill",
        lambda pid, sent: signals.append((pid, sent)),
    )

    with pytest.raises(ServiceStartError, match="could not start"):
        manager.start(tunnel=False)

    assert signals and signals[0][0] == 4321
    assert not manager.state_path.exists()


def test_ngrok_endpoint_conflict_is_converted_to_friendly_start_error(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    manager = _service_manager(tmp_path)
    monkeypatch.setattr(manager, "status", lambda: {"running": False})

    def popen(*args: object, **kwargs: object) -> SimpleNamespace:
        manager.log_path.parent.mkdir(parents=True, exist_ok=True)
        manager.log_path.write_text("ngrok failed with ERR_NGROK_334\n", encoding="utf-8")
        return SimpleNamespace(pid=4321)

    monkeypatch.setattr("research_gateway.operations.service.subprocess.Popen", popen)
    monkeypatch.setattr(
        manager,
        "_wait_for_health",
        lambda pid: (_ for _ in ()).throw(RuntimeError("process exited")),
    )
    monkeypatch.setattr(manager, "_process_matches", lambda pid: False)

    with pytest.raises(ServiceStartError, match="ngrok endpoint is already online"):
        manager.start(tunnel=True)

    assert not manager.state_path.exists()


def test_service_stop_and_restart_use_validated_pid(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    manager = _service_manager(tmp_path)
    manager.runtime_directory.mkdir(parents=True)
    manager.state_path.write_text('{"pid": 4321}', encoding="utf-8")
    matches = iter([True, False, False])
    monkeypatch.setattr(manager, "_process_matches", lambda pid: next(matches))
    signals: list[tuple[int, int]] = []
    monkeypatch.setattr(
        "research_gateway.operations.service.os.kill",
        lambda pid, sent: signals.append((pid, sent)),
    )

    assert manager.stop(timeout_seconds=0) == {"running": False, "stopped": True, "pid": 4321}
    assert signals[0][0] == 4321
    assert not manager.state_path.exists()

    monkeypatch.setattr(manager, "stop", lambda: {"stopped": True})
    monkeypatch.setattr(manager, "start", lambda *, tunnel: {"running": True, "tunnel": tunnel})
    assert manager.restart(tunnel=True) == {"running": True, "tunnel": True}


def test_service_status_reports_safe_local_and_public_locations(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    manager = _service_manager(tmp_path)
    manager.runtime_directory.mkdir(parents=True)
    manager.state_path.write_text(
        '{"pid": 4321, "started_at": "2026-08-18T00:00:00Z"}', encoding="utf-8"
    )
    manager.tunnel_state_path.write_text(
        '{"public_url": "https://safe.ngrok.app"}', encoding="utf-8"
    )
    monkeypatch.setattr(manager, "_process_matches", lambda pid: True)
    monkeypatch.setattr(manager, "_health_ok", lambda: True)

    result = manager.status()

    assert result["running"] is True
    assert result["classification"] == "managed"
    assert result["local_ui_url"] == "http://127.0.0.1:8877/ui"
    assert result["local_mcp_url"] == "http://127.0.0.1:8877/mcp"
    assert result["public_mcp_url"] == "https://safe.ngrok.app/mcp"
    assert result["database_path"].endswith("research_gateway.db")


def test_service_helpers_reject_stale_state_and_validate_health(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    manager = _service_manager(tmp_path)
    manager.runtime_directory.mkdir(parents=True)
    manager.state_path.write_text("not-json", encoding="utf-8")
    assert manager._read_state() == {}
    assert manager._process_matches(1) is False

    class Response(BytesIO):
        status = 200

        def __enter__(self) -> Response:
            return self

        def __exit__(self, *args: object) -> None:
            return None

    monkeypatch.setattr(
        "research_gateway.operations.service.urllib.request.urlopen",
        lambda url, timeout: Response(b'{"status":"ok","service":"research-gateway"}'),
    )
    assert manager._health_ok() is True

    monkeypatch.setattr(
        "research_gateway.operations.service.urllib.request.urlopen",
        lambda url, timeout: Response(b'{"status":"ok","service":"something-else"}'),
    )
    assert manager._health_ok() is False

    manager.state_path.write_text('{"pid": 123}', encoding="utf-8")
    monkeypatch.setattr(manager, "_process_matches", lambda pid: False)
    monkeypatch.setattr(manager, "_health_ok", lambda: False)
    monkeypatch.setattr(manager, "_port_in_use", lambda: False)
    assert manager.status()["running"] is False
    assert manager.state_path.exists()


def test_service_cli_reports_expected_conflict_without_traceback(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    config = tmp_path / "config.toml"
    config.write_text(
        "\n".join(
            (
                "[service]",
                'host = "127.0.0.1"',
                "port = 8877",
                "[database]",
                f'path = "{tmp_path / "gateway.db"}"',
                "[runtime]",
                f'directory = "{tmp_path / "runtime"}"',
            )
        ),
        encoding="utf-8",
    )
    monkeypatch.setattr(
        ServiceManager,
        "start",
        lambda self, *, tunnel: (_ for _ in ()).throw(
            ServiceStartError("Port 8877 is already in use by another service.")
        ),
    )

    result = CliRunner().invoke(app, ["service", "start", "--tunnel", "--config", str(config)])

    assert result.exit_code == 2
    assert "Port 8877 is already in use by another service." in result.output
    assert "Traceback" not in result.output


def test_service_cli_start_is_friendly_noop_for_existing_gateway(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    config = tmp_path / "config.toml"
    config.write_text(
        "\n".join(
            (
                "[service]",
                'host = "127.0.0.1"',
                "port = 8877",
                "[database]",
                f'path = "{tmp_path / "gateway.db"}"',
            )
        ),
        encoding="utf-8",
    )
    monkeypatch.setattr(
        ServiceManager,
        "start",
        lambda self, *, tunnel: {
            "running": True,
            "classification": "unmanaged",
            "pid": 7654,
            "local_ui_url": "http://127.0.0.1:8877/ui",
            "local_mcp_url": "http://127.0.0.1:8877/mcp",
            "public_mcp_url": "https://example.ngrok.app/mcp",
            "started": False,
        },
    )

    result = CliRunner().invoke(app, ["service", "start", "--tunnel", "--config", str(config)])

    assert result.exit_code == 0
    assert "Research Gateway is already running." in result.output
    assert "Service: running (existing/unmanaged instance)" in result.output
    assert "Public MCP: https://example.ngrok.app/mcp" in result.output
    assert "Service start: no action required." in result.output


def test_service_cli_uses_matching_supervisor_for_start(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    config = tmp_path / "config.toml"
    config.write_text(
        "\n".join(
            (
                "[service]",
                'host = "127.0.0.1"',
                "port = 8877",
                "[database]",
                f'path = "{tmp_path / "gateway.db"}"',
            )
        ),
        encoding="utf-8",
    )
    monkeypatch.setattr(
        SystemdUserSupervisor,
        "manages_config",
        property(lambda self: True),
    )
    monkeypatch.setattr(
        SystemdUserSupervisor,
        "start",
        lambda self, manager: {
            "running": True,
            "classification": "supervised",
            "pid": 4321,
            "supervisor_installed": True,
            "supervisor_enabled": True,
            "supervisor_restarts": 1,
            "supervisor_log_path": str(tmp_path / "supervisor.log"),
            "started": True,
        },
    )
    monkeypatch.setattr(
        ServiceManager,
        "start",
        lambda *args, **kwargs: pytest.fail("detached start must not be used"),
    )

    result = CliRunner().invoke(app, ["service", "start", "--config", str(config)])

    assert result.exit_code == 0
    assert "Service: running (supervised with automatic recovery)" in result.output
    assert "Supervisor: enabled" in result.output
    assert "Automatic restarts: 1" in result.output


def test_service_install_replaces_owned_detached_process_and_starts_supervisor(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    config = tmp_path / "config.toml"
    config.write_text(
        "\n".join(
            (
                "[service]",
                'host = "127.0.0.1"',
                "port = 8877",
                "[database]",
                f'path = "{tmp_path / "gateway.db"}"',
            )
        ),
        encoding="utf-8",
    )
    calls: list[object] = []
    monkeypatch.setattr(
        SystemdUserSupervisor,
        "validate_durable_location",
        lambda self: calls.append("validate"),
    )
    monkeypatch.setattr(
        SystemdUserSupervisor,
        "installed",
        property(lambda self: False),
    )
    monkeypatch.setattr(
        ServiceManager,
        "status",
        lambda self: {"running": True, "classification": "managed", "pid": 111},
    )
    monkeypatch.setattr(
        ServiceManager,
        "stop",
        lambda self: calls.append("stop-detached") or {"stopped": True},
    )
    monkeypatch.setattr(
        SystemdUserSupervisor,
        "install",
        lambda self, *, tunnel: calls.append(("install", tunnel)) or {"installed": True},
    )
    monkeypatch.setattr(
        SystemdUserSupervisor,
        "start",
        lambda self, manager: (
            calls.append("start-supervisor")
            or {
                "running": True,
                "classification": "supervised",
                "supervisor_enabled": True,
                "started": True,
            }
        ),
    )

    result = CliRunner().invoke(
        app,
        [
            "service",
            "install",
            "--tunnel",
            "--config",
            str(config),
            "--working-directory",
            str(tmp_path / "repository"),
            "--python-executable",
            str(tmp_path / ".venv" / "bin" / "python"),
        ],
    )

    assert result.exit_code == 0
    assert calls == ["validate", "stop-detached", ("install", True), "start-supervisor"]
    assert "Supervision installed and service started." in result.output
