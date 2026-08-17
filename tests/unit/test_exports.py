from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from research_gateway.db.database import EvidenceDatabase
from research_gateway.domain.models import SourceRecord
from research_gateway.services.exports import ExportService


@pytest.mark.asyncio
async def test_all_export_formats_include_evidence_and_provenance(tmp_path: Path) -> None:
    database = EvidenceDatabase(tmp_path / "evidence.db")
    await database.migrate()
    await database.create_study("s1", "Study", "Purpose")
    run = await database.create_search_run(
        study_id="s1",
        topic_id=None,
        provider="arxiv",
        mode="save",
        label="baseline",
        search_intent="Find baseline",
        provider_query="all:gateway",
        filters={},
        sort={},
        requested_limit=1,
    )
    outcome = await database.ingest_search_hit(
        run.search_run_id,
        1,
        SourceRecord(
            provider="arxiv",
            provider_record_id="2401.1",
            title="Research gateway evidence",
            authors=[{"name": "Ada Example"}],
            abstract="A result",
            year=2024,
            identifiers={"arxiv_id": "2401.1"},
        ),
    )
    await database.complete_search_run(
        run.search_run_id,
        provider_reported_total=1,
        retrieved_count=1,
        complete=True,
        pagination={},
        new_evidence_count=1,
    )
    await database.set_screening(
        outcome.evidence_id, "final", reason=None, note="accepted", actor="test"
    )
    exporter = ExportService(database)
    for format in ("json", "csv", "xlsx", "markdown"):
        result = await exporter.export(tmp_path / f"export.{format}", format=format, study_id="s1")
        assert result["evidence_count"] == 1
        assert result["discovery_count"] == 1
        assert Path(result["path"]).is_file()
    payload = json.loads((tmp_path / "export.json").read_text())
    assert payload["discoveries"][0]["provider"] == "arxiv"
    assert payload["search_runs"][0]["provider_query"] == "all:gateway"
    assert "Research gateway evidence" in (tmp_path / "export.csv").read_text()
    assert "Research gateway evidence" in (tmp_path / "export.markdown").read_text()
    workbook = load_workbook(tmp_path / "export.xlsx", read_only=True)
    assert workbook.sheetnames == [
        "Evidence",
        "Discoveries",
        "Search Runs",
        "Screening",
        "Topics",
    ]
