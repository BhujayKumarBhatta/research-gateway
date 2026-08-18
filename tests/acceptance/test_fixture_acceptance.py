from __future__ import annotations

import base64
import csv
import json
from pathlib import Path
from typing import Any

import httpx
import pytest
from mcp.client import Client
from openpyxl import load_workbook

from research_gateway.config import (
    AclSettings,
    ArxivSettings,
    GithubSettings,
    ScopusSettings,
    Settings,
    ZoteroSettings,
)
from research_gateway.db.database import EvidenceDatabase
from research_gateway.integrations.github import GithubAdapter
from research_gateway.integrations.zotero import ZoteroAdapter
from research_gateway.mcp.server import create_mcp_server
from research_gateway.runtime import GatewayRuntime
from research_gateway.services.exports import ExportService
from research_gateway.services.research import ResearchService
from research_gateway.sources.acl_anthology import AclAnthologyAdapter
from research_gateway.sources.arxiv import ArxivAdapter
from research_gateway.sources.registry import SourceRegistry
from research_gateway.sources.scopus import ScopusAdapter
from research_gateway.tunnel import NgrokTunnel

FIXTURES = Path(__file__).parents[1] / "fixtures"
SECRETS = (
    "fixture-scopus-secret",
    "fixture-zotero-secret",
    "fixture-github-secret",
    "fixture-ngrok-secret",
    "fixture-remote-mcp-secret",
)


class _Listener:
    def url(self) -> str:
        return "https://fixture.ngrok.app"


class _NgrokBackend:
    def __init__(self) -> None:
        self.forwarded: tuple[object, dict[str, object]] | None = None
        self.disconnected: str | None = None

    def forward(self, addr: object = None, **options: object) -> _Listener:
        self.forwarded = (addr, options)
        return _Listener()

    def disconnect(self, url: str | None = None) -> None:
        self.disconnected = url


def _scopus_payload() -> dict[str, Any]:
    payload = json.loads((FIXTURES / "scopus_search.json").read_text(encoding="utf-8"))
    entries = payload["search-results"]["entry"]
    entries.append(
        {
            "eid": "2-s2.0-125",
            "dc:title": "Distinct Scopus Failure Taxonomy",
            "dc:creator": "Turing, Alan",
            "prism:coverDate": "2023-04-03",
            "prism:publicationName": "Systems Evidence",
            "subtype": "ar",
        }
    )
    payload["search-results"]["opensearch:totalResults"] = "42"
    payload["search-results"]["opensearch:itemsPerPage"] = "3"
    return payload


def _arxiv_payload() -> bytes:
    entries = [
        (
            "2601.00001",
            "An arXiv version of the Scopus paper",
            "Ada Lovelace",
            "2025",
            "10.1000/test",
        ),
        ("2601.00002", "A Distinct Open Failure Study", "Barbara Liskov", "2026", None),
        ("2601.00003", "Reliable Language Models: Extended Study", "Grace Hopper", "2024", None),
    ]
    rendered = []
    for identifier, title, author, year, doi in entries:
        doi_xml = f"<arxiv:doi>{doi}</arxiv:doi>" if doi else ""
        rendered.append(
            f"""
  <entry>
    <id>http://arxiv.org/abs/{identifier}v1</id>
    <updated>{year}-02-02T00:00:00Z</updated>
    <published>{year}-01-01T00:00:00Z</published>
    <title>{title}</title><summary>Fixture abstract.</summary>
    <author><name>{author}</name></author>{doi_xml}
    <arxiv:primary_category term="cs.CL"/><category term="cs.CL"/>
    <link href="http://arxiv.org/abs/{identifier}v1" rel="alternate" type="text/html"/>
  </entry>"""
        )
    return (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<feed xmlns="http://www.w3.org/2005/Atom" '
        'xmlns:opensearch="http://a9.com/-/spec/opensearch/1.1/" '
        'xmlns:arxiv="http://arxiv.org/schemas/atom">'
        "<title>Fixture</title><opensearch:totalResults>3</opensearch:totalResults>"
        "<opensearch:startIndex>0</opensearch:startIndex>"
        "<opensearch:itemsPerPage>3</opensearch:itemsPerPage>" + "".join(rendered) + "</feed>"
    ).encode()


@pytest.mark.asyncio
async def test_complete_multi_source_fixture_acceptance(tmp_path: Path) -> None:
    scopus_payload = _scopus_payload()
    arxiv_payload = _arxiv_payload()
    remote_calls: list[tuple[str, str, Any]] = []
    zotero_item_calls = 0
    zotero_items: dict[str, dict[str, Any]] = {
        "SEARCH1": {
            "key": "SEARCH1",
            "version": 7,
            "data": {
                "itemType": "journalArticle",
                "title": "A Fine-Tuning Search Result",
                "date": "2025",
                "url": "https://example.test/fine-tuning",
                "creators": [{"creatorType": "author", "name": "A. Researcher"}],
                "collections": [],
            },
        }
    }

    async def scopus_handler(request: httpx.Request) -> httpx.Response:
        assert request.headers["X-ELS-APIKey"] == SECRETS[0]
        return httpx.Response(200, json=scopus_payload)

    async def arxiv_handler(_: httpx.Request) -> httpx.Response:
        return httpx.Response(200, content=arxiv_payload)

    async def zotero_handler(request: httpx.Request) -> httpx.Response:
        nonlocal zotero_item_calls
        assert request.headers["Zotero-API-Key"] == SECRETS[1]
        remote_calls.append((request.method, request.url.path, None))
        if request.method == "GET" and request.url.path == "/keys/current":
            return httpx.Response(
                200,
                json={
                    "userID": 42,
                    "access": {"user": {"library": True, "write": True, "notes": True}},
                },
            )
        if request.method == "GET" and request.url.path.endswith("/collections"):
            return httpx.Response(200, json=[])
        if request.method == "GET" and "/items/" in request.url.path:
            item_key = request.url.path.rsplit("/", 1)[-1]
            return httpx.Response(200, json=zotero_items[item_key])
        if request.method == "GET" and request.url.path.endswith("/items"):
            if request.url.params.get("q") == "fine-tuning":
                return httpx.Response(200, json=[zotero_items["SEARCH1"]])
            return httpx.Response(200, json=[])
        if request.method == "POST" and request.url.path.endswith("/collections"):
            return httpx.Response(200, json={"successful": {"0": {"key": "FINAL"}}})
        if request.method == "POST" and request.url.path.endswith("/items"):
            zotero_item_calls += 1
            items = json.loads(request.content)
            assert len(items) == 1
            assert items[0]["collections"] == ["FINAL"]
            zotero_items["ITEM1"] = {
                "key": "ITEM1",
                "version": 8,
                "data": items[0],
            }
            return httpx.Response(
                200,
                json={"successful": {"0": zotero_items["ITEM1"]}, "failed": {}},
            )
        raise AssertionError(request.url.path)

    async def github_handler(request: httpx.Request) -> httpx.Response:
        assert request.headers["Authorization"] == f"Bearer {SECRETS[2]}"
        body = json.loads(request.content) if request.content else None
        remote_calls.append((request.method, request.url.path, body))
        path = request.url.path
        if path.endswith("/contents/README.md"):
            return httpx.Response(
                200,
                json={
                    "type": "file",
                    "path": "README.md",
                    "sha": "readme-sha",
                    "size": 16,
                    "encoding": "base64",
                    "content": base64.b64encode(b"fixture contents").decode(),
                },
            )
        if path == "/repos/example/research" and request.method == "GET":
            return httpx.Response(200, json={"default_branch": "trunk"})
        if path.endswith("/issues") and request.method == "POST":
            return httpx.Response(201, json={"number": 5, "html_url": "https://github.test/5"})
        if path.endswith("/git/ref/heads/trunk"):
            return httpx.Response(200, json={"object": {"sha": "base-sha"}})
        if path.endswith("/git/refs"):
            return httpx.Response(201, json={"ref": "refs/heads/research/fixture"})
        if path.endswith("/git/commits/base-sha"):
            return httpx.Response(200, json={"tree": {"sha": "tree-base"}})
        if path.endswith("/git/blobs"):
            return httpx.Response(201, json={"sha": "blob-sha"})
        if path.endswith("/git/trees"):
            return httpx.Response(201, json={"sha": "tree-new"})
        if path.endswith("/git/commits"):
            return httpx.Response(201, json={"sha": "commit-new"})
        if path.endswith("/git/refs/heads/research/fixture"):
            return httpx.Response(200, json={"ref": "updated"})
        if path.endswith("/pulls"):
            return httpx.Response(201, json={"number": 7, "html_url": "https://github.test/7"})
        raise AssertionError(path)

    clients = [
        httpx.AsyncClient(transport=httpx.MockTransport(scopus_handler)),
        httpx.AsyncClient(transport=httpx.MockTransport(arxiv_handler)),
        httpx.AsyncClient(transport=httpx.MockTransport(zotero_handler)),
        httpx.AsyncClient(transport=httpx.MockTransport(github_handler)),
    ]
    settings = Settings.model_validate(
        {
            "database": {"path": tmp_path / "gateway.db"},
            "scopus": {"api_key": SECRETS[0]},
            "acl_anthology": {"index_path": FIXTURES / "acl_index.json"},
            "arxiv": {"polite_delay_seconds": 0},
            "zotero": {
                "enabled": True,
                "api_key": SECRETS[1],
                "library_id": "42",
                "collection_name": "Selected evidence",
            },
            "github": {"enabled": True, "token": SECRETS[2]},
            "tunnel": {"authtoken": SECRETS[3]},
            "mcp_remote_auth": {"token": SECRETS[4]},
        }
    )
    database = EvidenceDatabase(settings.database.path)
    sources = SourceRegistry(
        [
            ScopusAdapter(ScopusSettings(api_key=SECRETS[0]), client=clients[0]),
            AclAnthologyAdapter(AclSettings(index_path=FIXTURES / "acl_index.json")),
            ArxivAdapter(ArxivSettings(polite_delay_seconds=0), client=clients[1]),
        ]
    )
    runtime = GatewayRuntime(
        settings=settings,
        database=database,
        sources=sources,
        research=ResearchService(database, sources),
        exports=ExportService(database),
        zotero=ZoteroAdapter(
            ZoteroSettings(
                enabled=True,
                api_key=SECRETS[1],
                library_id="42",
                collection_name="Selected evidence",
            ),
            database,
            client=clients[2],
        ),
        github=GithubAdapter(
            GithubSettings(enabled=True, token=SECRETS[2]), database, client=clients[3]
        ),
    )
    safe_outputs: list[object] = []
    await runtime.start()
    try:
        async with Client(create_mcp_server(runtime)) as client:
            listed = await client.list_tools()
            assert {
                "gateway_status",
                "source_list",
                "research_explore_search",
                "zotero_search",
                "zotero_get_item",
                "zotero_credential_status",
            } <= {tool.name for tool in listed.tools}
            status = await client.call_tool("gateway_status")
            sources_status = await client.call_tool("source_list")
            zotero_permissions = await client.call_tool("zotero_credential_status")
            zotero_search = await client.call_tool("zotero_search", {"query": "fine-tuning"})
            searched_item = zotero_search.structured_content["items"][0]
            assert searched_item["item_key"] == "SEARCH1"
            zotero_item = await client.call_tool(
                "zotero_get_item", {"item_key": searched_item["item_key"]}
            )
            assert zotero_item.structured_content["item_key"] == searched_item["item_key"]
            assert zotero_item.structured_content["title"] == searched_item["title"]
            assert status.structured_content["database_schema"] == 2
            assert zotero_permissions.structured_content["library_read"] is True
            assert zotero_permissions.structured_content["library_write"] is True
            assert all(
                source["available"]
                for source in sources_status.structured_content["sources"]
                if source["name"] in {"scopus", "acl_anthology", "arxiv"}
            )
            safe_outputs.extend(
                [
                    status.structured_content,
                    sources_status.structured_content,
                    zotero_permissions.structured_content,
                    zotero_search.structured_content,
                    zotero_item.structured_content,
                ]
            )

            await client.call_tool(
                "study_create",
                {
                    "study_id": "acceptance-study",
                    "name": "Acceptance study",
                    "system_test": True,
                },
            )
            await client.call_tool(
                "topic_create",
                {
                    "study_id": "acceptance-study",
                    "topic_id": "failure-types",
                    "name": "Failure types",
                },
            )
            exact_scopus_query = "TITLE-ABS-KEY(failure)"
            explored = await client.call_tool(
                "research_explore_search",
                {
                    "study_id": "acceptance-study",
                    "topic_id": "failure-types",
                    "provider": "scopus",
                    "search_intent": "Tune the licensed query",
                    "provider_query": exact_scopus_query,
                },
            )
            assert explored.structured_content["provider_reported_total"] == 42
            assert await database.count_rows("search_runs") == 1
            assert await database.count_rows("search_hits") == 0
            assert await database.count_rows("evidence") == 0

            save_results = []
            for provider, query, limit, intent in (
                ("scopus", exact_scopus_query, 3, "Capture licensed failure evidence"),
                ("acl_anthology", "", 2, "Capture official ACL evidence"),
                ("arxiv", "all:failure", 3, "Capture open preprints"),
            ):
                result = await client.call_tool(
                    "research_save_search",
                    {
                        "study_id": "acceptance-study",
                        "topic_id": "failure-types",
                        "provider": provider,
                        "search_intent": intent,
                        "provider_query": query,
                        "max_records": limit,
                    },
                )
                assert result.is_error is False
                save_results.append(result.structured_content)
            assert [result["retrieved_count"] for result in save_results] == [3, 2, 3]
            assert await database.count_rows("search_runs") == 4
            assert await database.count_rows("search_hits") == 8
            assert await database.count_rows("evidence") == 6
            duplicates = await client.call_tool("evidence_possible_duplicates")
            assert duplicates.structured_content["possible_duplicates"]

            runs = await client.call_tool(
                "search_runs_list", {"study_id": "acceptance-study", "topic_id": "failure-types"}
            )
            assert len(runs.structured_content["search_runs"]) == 4
            for stored in runs.structured_content["search_runs"]:
                assert stored["provider_query"] is not None
                assert stored["search_intent"]
                assert stored["executed_at_utc"]
                detailed = await client.call_tool(
                    "search_run_get", {"search_run_id": stored["search_run_id"]}
                )
                assert detailed.structured_content["study_id"] == "acceptance-study"
                assert detailed.structured_content["topic_id"] == "failure-types"
                assert detailed.structured_content["status"] == "completed"
            assert save_results[0]["provider_reported_total"] == 42

            evidence = await client.call_tool(
                "evidence_search", {"study_id": "acceptance-study", "limit": 20}
            )
            items = evidence.structured_content["items"]
            assert len(items) == 6
            assert any(item["review_status"] == "preprint" for item in items)
            assert any(item["publication_type"] == "journal_article" for item in items)
            decisions = (
                (items[0]["evidence_id"], "excluded", "Outside the review scope"),
                (items[1]["evidence_id"], "included", None),
                (items[2]["evidence_id"], "final", None),
                (items[3]["evidence_id"], "candidate", None),
            )
            for evidence_id, decision, reason in decisions:
                screened = await client.call_tool(
                    "evidence_set_screening",
                    {
                        "evidence_id": evidence_id,
                        "status": decision,
                        "reason": reason,
                        "note": f"Fixture {decision}",
                        "actor": "acceptance",
                    },
                )
                assert screened.structured_content["history"][-1]["new_status"] == decision
            final_evidence_id = items[2]["evidence_id"]
            detailed_evidence = await client.call_tool(
                "evidence_get", {"evidence_id": items[0]["evidence_id"]}
            )
            assert detailed_evidence.structured_content["discoveries"]

            xlsx_path = tmp_path / "evidence.xlsx"
            csv_path = tmp_path / "evidence.csv"
            excel = await client.call_tool(
                "evidence_export_excel",
                {"path": str(xlsx_path), "study_id": "acceptance-study"},
            )
            exported_csv = await client.call_tool(
                "evidence_export_csv",
                {"path": str(csv_path), "study_id": "acceptance-study"},
            )
            assert excel.structured_content["evidence_count"] == 6
            assert exported_csv.structured_content["evidence_count"] == 6
            workbook = load_workbook(xlsx_path, read_only=True)
            assert workbook.sheetnames == [
                "Evidence",
                "Discoveries",
                "Search Runs",
                "Screening",
                "Topics",
            ]
            with csv_path.open(encoding="utf-8", newline="") as stream:
                exported_rows = list(csv.DictReader(stream))
            assert len(exported_rows) == 6
            assert {"evidence_code", "screening_status", "first_discovery"} <= exported_rows[
                0
            ].keys()

            calls_before_dry_run = len(remote_calls)
            zotero_plan = await client.call_tool(
                "zotero_sync_corpus", {"study_id": "acceptance-study"}
            )
            assert zotero_plan.structured_content["would_create"] == 1
            assert len(remote_calls) > calls_before_dry_run
            assert all(method == "GET" for method, _, _ in remote_calls[calls_before_dry_run:])
            zotero_first = await client.call_tool(
                "zotero_sync_corpus", {"study_id": "acceptance-study", "dry_run": False}
            )
            zotero_second = await client.call_tool(
                "zotero_sync_corpus", {"study_id": "acceptance-study", "dry_run": False}
            )
            assert zotero_first.structured_content["created"] == 1
            created_item = zotero_first.structured_content["created_items"][0]
            assert created_item["evidence_id"] == final_evidence_id
            assert created_item["item_key"] == "ITEM1"
            created_readback = await client.call_tool(
                "zotero_get_item", {"item_key": created_item["item_key"]}
            )
            assert created_readback.structured_content["item_key"] == "ITEM1"
            assert created_readback.structured_content["title"] == created_item["title"]
            assert zotero_second.structured_content["created"] == 0
            assert zotero_second.structured_content["already_linked"] == 1
            assert zotero_item_calls == 1
            assert await database.get_zotero_link(final_evidence_id, "user", "42")

            github_file = await client.call_tool(
                "github_read_file", {"repository": "example/research", "path": "README.md"}
            )
            assert github_file.structured_content["content"] == "fixture contents"
            issue = await client.call_tool(
                "github_create_issue",
                {
                    "repository": "example/research",
                    "title": "Review fixture evidence",
                    "body": "Please review.",
                    "dry_run": False,
                },
            )
            assert issue.structured_content["number"] == 5
            change = {
                "repository": "example/research",
                "branch": "research/fixture",
                "files": {"reports/evidence.md": "Safe fixture evidence"},
                "commit_message": "Add fixture evidence",
                "pull_request_title": "Add fixture evidence",
                "pull_request_body": "Review this fixture export.",
            }
            github_plan = await client.call_tool("github_propose_change", change)
            github_real = await client.call_tool(
                "github_propose_change", {**change, "dry_run": False}
            )
            assert github_plan.structured_content["dry_run"] is True
            assert github_real.structured_content["commit_sha"] == "commit-new"
            assert github_real.structured_content["pull_request_number"] == 7
            assert not any(
                method == "PATCH" and path.endswith("/git/refs/heads/trunk")
                for method, path, _ in remote_calls
            )

            audit = await client.call_tool("audit_recent", {"limit": 100})
            operations = {event["operation"] for event in audit.structured_content["events"]}
            assert {
                "search.explore",
                "search.save",
                "export.create",
                "zotero.sync_final_corpus",
            } <= operations
            safe_outputs.extend(
                [
                    explored.structured_content,
                    *save_results,
                    runs.structured_content,
                    evidence.structured_content,
                    detailed_evidence.structured_content,
                    duplicates.structured_content,
                    excel.structured_content,
                    exported_csv.structured_content,
                    zotero_plan.structured_content,
                    zotero_first.structured_content,
                    created_readback.structured_content,
                    github_file.structured_content,
                    issue.structured_content,
                    github_plan.structured_content,
                    github_real.structured_content,
                    audit.structured_content,
                ]
            )

        backend = _NgrokBackend()
        tunnel_state = tmp_path / "runtime" / "tunnel.json"
        tunnel = NgrokTunnel(settings, backend=backend, state_path=tunnel_state)
        public = await tunnel.astart()
        assert public.exposed_paths == ["/health", "/mcp"]
        assert public.ui_exposed is False
        assert tunnel_state.is_file()
        await tunnel.astop()
        assert backend.disconnected == "https://fixture.ngrok.app"
        assert not tunnel_state.exists()

        serialized_outputs = json.dumps(safe_outputs, default=str)
        audit_bytes = settings.database.path.read_bytes()
        csv_text = csv_path.read_text(encoding="utf-8")
        ui_text = "".join(
            path.read_text(encoding="utf-8")
            for path in (Path(__file__).parents[2] / "ui").glob("*.html")
        )
        for secret in SECRETS:
            encoded = secret.encode()
            assert secret not in serialized_outputs
            assert encoded not in audit_bytes
            assert secret not in csv_text
            assert encoded not in xlsx_path.read_bytes()
            assert secret not in ui_text
    finally:
        await runtime.aclose()
        for client in clients:
            await client.aclose()
